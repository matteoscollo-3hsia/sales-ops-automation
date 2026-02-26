from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook

import primer_ops.primer as primer


class _FakeResponses:
    def __init__(self, output_text: str) -> None:
        self._output_text = output_text

    def create(self, **_kwargs):
        return {"output_text": self._output_text}


class _FakeOpenAI:
    def __init__(self, *args, **_kwargs) -> None:
        self.responses = _FakeResponses("This is the primer content.")


def _write_minimal_prompt_library(path: Path) -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Company Intro"

    ws["A1"] = "Instructions"
    ws["A2"] = "Web Search"
    ws["B2"] = "Disable"
    ws["A3"] = "Deep Research"
    ws["B3"] = "Disable"

    ws["A5"] = "Prompts"
    ws["A6"] = "Step 1 - Company Introduction"
    ws["A7"] = "Suggested Prompt"
    ws["B7"] = "Write intro about {company_name}."

    workbook.save(path)


def test_generate_primer_no_headings(tmp_path, monkeypatch) -> None:
    monkeypatch.setattr(primer, "OpenAI", _FakeOpenAI)
    monkeypatch.setattr(primer, "load_dotenv", lambda *args, **kwargs: None)
    monkeypatch.setattr(primer, "find_dotenv", lambda *args, **kwargs: "")

    prompt_path = tmp_path / "prompts.xlsx"
    _write_minimal_prompt_library(prompt_path)

    lead_path = tmp_path / "lead_input.json"
    lead_path.write_text(json.dumps({"company_name": "Acme Corp"}), encoding="utf-8")

    monkeypatch.setenv("PROMPT_LIBRARY_PATH", str(prompt_path))
    monkeypatch.delenv("INCLUDE_HEADINGS", raising=False)

    import primer_ops.render_docx as render_docx

    monkeypatch.setattr(render_docx, "render_primer_docx", lambda *args, **kwargs: None)
    monkeypatch.setattr(
        primer, "_resolve_template_path", lambda: tmp_path / "fake_template.docx"
    )

    output_dir = tmp_path / "out"
    primer.generate_primer(
        output_dir=str(output_dir),
        lead_input=str(lead_path),
        resume=False,
    )

    primer_text = (output_dir / "primer.md").read_text(encoding="utf-8")
    assert "Company Intro" not in primer_text
    assert "Company Introduction" not in primer_text
    assert "## " not in primer_text
    assert "### " not in primer_text
    assert primer_text.lstrip().startswith("This is the primer content.")

    sources = json.loads((output_dir / "sources.json").read_text(encoding="utf-8"))
    assert sources["sheets"][0]["name"] == "Company Intro"
    assert sources["sheets"][0]["steps"][0]["step_number"] == 1
    assert sources["sheets"][0]["steps"][0]["title"] == "Company Introduction"
