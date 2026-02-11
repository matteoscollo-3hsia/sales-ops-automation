from __future__ import annotations

import json
import os
from pathlib import Path
import re
import time
import traceback
from datetime import datetime
import uuid
from typing import Any, Iterable, Optional

from primer_ops.progress import spinner, format_seconds
from dotenv import find_dotenv, load_dotenv
from openai import APITimeoutError, OpenAI, NotFoundError, RateLimitError
from openpyxl import load_workbook

from primer_ops.client_repo import ensure_client_repo, sanitize_folder_name
from primer_ops.config import (
    get_include_headings,
    get_lead_input_path,
    get_output_base_dir,
    get_output_dir,
)

def _normalize(text: str) -> str:
    return " ".join(text.strip().split()).lower()


def _iter_cells(ws) -> Iterable[Any]:
    for row in ws.iter_rows():
        for cell in row:
            yield cell


def _find_anchor_exact(ws, label: str, start_row: int = 1) -> Optional[Any]:
    target = _normalize(label)
    for cell in _iter_cells(ws):
        if cell.row < start_row:
            continue
        if isinstance(cell.value, str) and _normalize(cell.value) == target:
            return cell
    return None


def _find_anchor_exact_in_window(
    ws, label: str, start_row: int, end_row: int
) -> Optional[Any]:
    if end_row <= start_row:
        return None
    target = _normalize(label)
    for row in ws.iter_rows(min_row=start_row, max_row=end_row - 1):
        for cell in row:
            if isinstance(cell.value, str) and _normalize(cell.value) == target:
                return cell
    return None


def _find_anchor_contains(ws, label: str, start_row: int = 1) -> Optional[Any]:
    target = _normalize(label)
    for cell in _iter_cells(ws):
        if cell.row < start_row:
            continue
        if isinstance(cell.value, str) and target in _normalize(cell.value):
            return cell
    return None


def _require_anchor(cell: Optional[Any], label: str) -> Any:
    if cell is None:
        raise SystemExit(f"ERROR: missing anchor: {label}")
    return cell


def _first_right_value(ws, row: int, col: int, limit: int = 6) -> Optional[Any]:
    for cc in range(col + 1, col + limit + 1):
        val = ws.cell(row=row, column=cc).value
        if val is not None and str(val).strip() != "":
            return val
    return None


def _replace_placeholders(text: str, lead: dict[str, Any]) -> str:
    updated = text
    for key in sorted(lead.keys(), key=lambda k: len(str(k)), reverse=True):
        raw_value = lead[key]
        value = str(raw_value)
        updated = updated.replace(f"{{{{{key}}}}}", value)
        updated = updated.replace(f"{{{key}}}", value)
        updated = updated.replace(f"#{key}#", value)
        updated = re.sub(
            rf"<<\s*{re.escape(str(key))}\s*>>",
            value,
            updated,
            flags=re.IGNORECASE,
        )
    return updated


_REQUEST_TIMEOUT_SECONDS = 30 * 60
_REMINDER_LINE_RE = re.compile(r"^\s*\(here\s+copy\s+and\s+paste.*\)\s*$", re.IGNORECASE)
_REMINDER_SENTENCE = "(here copy and paste introduction from 'company and industry intro' step 1)"
_CONTEXT_HEADER_RE = re.compile(r"^\s*###\s*context\b", re.IGNORECASE)
_SECTION_HEADER_RE = re.compile(r"^\s*###\s+", re.IGNORECASE)
_URL_RE = re.compile(r"https?://[^\s)>\"]+", re.IGNORECASE)


def _strip_human_reminders(text: str) -> str:
    if not text:
        return text
    lines = text.splitlines()
    filtered: list[str] = []
    for line in lines:
        if _REMINDER_LINE_RE.match(line):
            continue
        if _REMINDER_SENTENCE in line.lower():
            continue
        filtered.append(line)
    return "\n".join(filtered)


def _extract_company_name(lead: dict[str, Any]) -> str:
    for key in ("company_name", "client"):
        value = lead.get(key)
        if value is None:
            continue
        text = str(value).strip()
        if text:
            return text
    return "unknown_company"


def _extract_output_dir_override(lead: dict[str, Any]) -> Path | None:
    for key in ("client_output_dir", "output_dir"):
        value = lead.get(key)
        if isinstance(value, str) and value.strip():
            return Path(value.strip())
    return None


def resolve_lead_input_path(lead_input: str | None) -> Path:
    if lead_input and lead_input.strip():
        return Path(lead_input.strip())
    env_path = get_lead_input_path()
    if env_path is not None:
        return env_path
    return Path("lead_input.json")


def resolve_output_dir(output_dir: str | None, lead: dict[str, Any]) -> Path:
    if output_dir and output_dir.strip():
        return Path(output_dir.strip())
    override = _extract_output_dir_override(lead)
    if override is not None:
        return override
    base_dir = get_output_base_dir() or get_output_dir()
    if base_dir is None:
        raise SystemExit(
            "ERROR: OUTPUT_BASE_DIR is not set. Please set it in the .env file. "
            "(Legacy OUTPUT_DIR is also supported.)"
        )
    company_name = _extract_company_name(lead)
    folder_name = sanitize_folder_name(company_name) or "unknown_company"
    return base_dir / folder_name


def resolve_output_targets(output_dir: str | None, lead: dict[str, Any]) -> dict[str, Any]:
    company_name = _extract_company_name(lead)
    output_dir_override: Path | None = None
    if output_dir and output_dir.strip():
        output_dir_override = Path(output_dir.strip())
    else:
        output_dir_override = _extract_output_dir_override(lead)

    if output_dir_override is not None:
        output_dir_override.mkdir(parents=True, exist_ok=True)
        return {
            "output_dir": output_dir_override,
            "run_dir": None,
            "output_dirs": [output_dir_override],
            "repo_root": None,
            "latest_dir": None,
        }

    base_dir = get_output_base_dir() or get_output_dir()
    if base_dir is None:
        raise SystemExit(
            "ERROR: OUTPUT_BASE_DIR is not set. Please set it in the .env file. "
            "(Legacy OUTPUT_DIR is also supported.)"
        )
    repo = ensure_client_repo(base_dir, company_name)
    repo_root = repo["repo_root"]
    latest_dir = repo["latest_dir"]
    runs_dir = repo["runs_dir"]
    run_date = datetime.now().strftime("%Y-%m-%d")
    run_id = uuid.uuid4().hex[:8]
    run_output_dir_path = runs_dir / f"{run_date}_{run_id}"
    run_output_dir_path.mkdir(parents=True, exist_ok=True)
    return {
        "output_dir": latest_dir,
        "run_dir": run_output_dir_path,
        "output_dirs": [latest_dir, run_output_dir_path],
        "repo_root": repo_root,
        "latest_dir": latest_dir,
    }


def _build_prev_context_block(
    prev_sheet_name: str | None, prev_sheet_output_text: str | None
) -> tuple[str, str | None, int]:
    prev_output = (prev_sheet_output_text or "").strip()
    prev_output_chars = len(prev_output)
    if not prev_output:
        return "### CONTEXT\n\n(No previous sheet context available.)\n", None, prev_output_chars
    prev_name_used = prev_sheet_name or "N/A"
    block = "### CONTEXT\n\n" + prev_output + "\n"
    return block, prev_name_used, prev_output_chars


def _post_process_prompt(
    prompt_original: str, prev_sheet_name: str | None, prev_sheet_output_text: str
) -> tuple[str, bool, str | None, int]:
    cleaned = _strip_human_reminders(prompt_original or "")
    context_block, prev_name_used, prev_output_chars = _build_prev_context_block(
        prev_sheet_name, prev_sheet_output_text
    )
    pattern = r"^###\s*CONTEXT\s*$.*?(?=^###\s+|\Z)"
    if re.search(pattern, cleaned, flags=re.MULTILINE | re.DOTALL):
        prompt_final = re.sub(pattern, context_block, cleaned, flags=re.MULTILINE | re.DOTALL)
    else:
        prompt_final = f"{context_block}\n{cleaned}"
    return prompt_final, True, prev_name_used, prev_output_chars


def _extract_output_text_from_response(response: Any) -> str | None:
    if response is None:
        return None
    if isinstance(response, dict):
        output_text = response.get("output_text")
        if isinstance(output_text, str) and output_text.strip():
            return output_text
        for item in response.get("output", []) or []:
            text = _extract_output_text_from_item(item)
            if text:
                return text
        return None
    output_text = getattr(response, "output_text", None)
    if isinstance(output_text, str) and output_text.strip():
        return output_text
    output_items = getattr(response, "output", None)
    if isinstance(output_items, list):
        for item in output_items:
            text = _extract_output_text_from_item(item)
            if text:
                return text
    return None


def _extract_output_text_from_item(item: Any) -> str | None:
    if item is None:
        return None
    if isinstance(item, dict):
        item_type = item.get("type")
        if item_type == "output_text":
            text = item.get("text")
            if isinstance(text, str) and text.strip():
                return text
        content = item.get("content")
        if isinstance(content, list):
            for part in content:
                text = _extract_output_text_from_item(part)
                if text:
                    return text
    else:
        item_type = getattr(item, "type", None)
        if item_type == "output_text":
            text = getattr(item, "text", None)
            if isinstance(text, str) and text.strip():
                return text
        content = getattr(item, "content", None)
        if isinstance(content, list):
            for part in content:
                text = _extract_output_text_from_item(part)
                if text:
                    return text
    return None


def _extract_urls_from_text(text: str) -> list[str]:
    if not text:
        return []
    urls: list[str] = []
    for match in _URL_RE.findall(text):
        cleaned = match.rstrip(").,]")
        if cleaned:
            urls.append(cleaned)
    return urls


def _extract_citations_from_response(
    response: Any, output_text: str | None = None
) -> list[str]:
    urls: list[str] = []

    def add_url(value: Any) -> None:
        if isinstance(value, str):
            cleaned = value.strip()
            if cleaned:
                urls.append(cleaned)

    def walk(obj: Any) -> None:
        if obj is None:
            return
        if isinstance(obj, dict):
            add_url(obj.get("url"))
            add_url(obj.get("source_url"))
            add_url(obj.get("source"))
            annotations = obj.get("annotations")
            if isinstance(annotations, list):
                for item in annotations:
                    walk(item)
            content = obj.get("content")
            if isinstance(content, list):
                for item in content:
                    walk(item)
        else:
            add_url(getattr(obj, "url", None))
            add_url(getattr(obj, "source_url", None))
            add_url(getattr(obj, "source", None))
            annotations = getattr(obj, "annotations", None)
            if isinstance(annotations, list):
                for item in annotations:
                    walk(item)
            content = getattr(obj, "content", None)
            if isinstance(content, list):
                for item in content:
                    walk(item)

    output_items = None
    if isinstance(response, dict):
        output_items = response.get("output")
    else:
        output_items = getattr(response, "output", None)
    if isinstance(output_items, list):
        for item in output_items:
            walk(item)
    if output_text:
        urls.extend(_extract_urls_from_text(output_text))

    seen: set[str] = set()
    deduped: list[str] = []
    for url in urls:
        if url in seen:
            continue
        seen.add(url)
        deduped.append(url)
    return deduped


def _ensure_response_text(step_entry: dict[str, Any]) -> str | None:
    existing = step_entry.get("response_text")
    if isinstance(existing, str) and existing.strip():
        return existing
    legacy = step_entry.get("output_text")
    if isinstance(legacy, str) and legacy.strip():
        step_entry["response_text"] = legacy
        return legacy
    derived = _extract_output_text_from_response(step_entry.get("response"))
    if isinstance(derived, str) and derived.strip():
        step_entry["response_text"] = derived
        return derived
    return None


def _coerce_str(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        return value
    return str(value)


def _coerce_int(value: Any) -> int | None:
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    try:
        return int(str(value))
    except (TypeError, ValueError):
        return None


def _coerce_bool(value: Any, default: bool = False) -> bool:
    if isinstance(value, bool):
        return value
    return default


def _coerce_str_list(value: Any) -> list[str]:
    if not value:
        return []
    if isinstance(value, str):
        return [value.strip()] if value.strip() else []
    if isinstance(value, list):
        items: list[str] = []
        for item in value:
            if isinstance(item, str) and item.strip():
                items.append(item.strip())
        return items
    return []


def _sanitize_sources_payload(payload: dict[str, Any]) -> dict[str, Any]:
    if not isinstance(payload, dict):
        return {"prompt_library_path": "", "sheets": []}
    sanitized: dict[str, Any] = {
        "prompt_library_path": str(payload.get("prompt_library_path", "") or ""),
        "sheets": [],
    }
    sheets = payload.get("sheets")
    if not isinstance(sheets, list):
        return sanitized
    for sheet in sheets:
        if not isinstance(sheet, dict):
            continue
        name = sheet.get("name")
        if not isinstance(name, str):
            continue
        sanitized_sheet: dict[str, Any] = {
            "name": name,
            "web_search": _coerce_bool(sheet.get("web_search"), False),
            "deep_research_requested": _coerce_bool(sheet.get("deep_research_requested"), False),
            "deep_research_effective": _coerce_bool(sheet.get("deep_research_effective"), False),
            "deep_research_error_reason": _coerce_str(sheet.get("deep_research_error_reason")),
            "steps": [],
        }
        steps = sheet.get("steps")
        if isinstance(steps, list):
            for step in steps:
                if not isinstance(step, dict):
                    continue
                step_number = _coerce_int(step.get("step_number"))
                if step_number is None:
                    continue
                prompt = _coerce_str(step.get("prompt"))
                if not prompt:
                    prompt = _coerce_str(step.get("prompt_final")) or _coerce_str(
                        step.get("prompt_original")
                    )
                response_text = _coerce_str(step.get("response_text"))
                if not response_text:
                    response_text = _coerce_str(step.get("output_text"))
                if not response_text:
                    response_text = _extract_output_text_from_response(step.get("response")) or ""
                error_value = step.get("error")
                if isinstance(error_value, dict):
                    error_value = error_value.get("message") or error_value.get("type")
                error_message = _coerce_str(error_value)
                citations = _coerce_str_list(step.get("citations") or step.get("urls"))
                sanitized_step: dict[str, Any] = {
                    "step_number": step_number,
                    "title": _coerce_str(step.get("title")) or f"Step {step_number}",
                    "prompt": prompt,
                    "response_text": response_text or "",
                    "model": _coerce_str(step.get("model")),
                    "reasoning_effort_requested": _coerce_str(
                        step.get("reasoning_effort_requested") or step.get("reasoning_effort")
                    ),
                    "reasoning_effort_effective": _coerce_str(step.get("reasoning_effort_effective")),
                    "web_search": _coerce_bool(step.get("web_search"), sanitized_sheet["web_search"]),
                    "deep_research_requested": _coerce_bool(
                        step.get("deep_research_requested"),
                        sanitized_sheet["deep_research_requested"],
                    ),
                    "deep_research_effective": _coerce_bool(step.get("deep_research_effective"), False),
                    "deep_research_error_reason": _coerce_str(step.get("deep_research_error_reason")),
                    "web_tool_type": _coerce_str(step.get("web_tool_type")),
                    "error": error_message,
                }
                if citations:
                    sanitized_step["citations"] = citations
                sanitized_sheet["steps"].append(sanitized_step)
        sanitized["sheets"].append(sanitized_sheet)
    return sanitized


def _model_supports_reasoning_effort(model: str | None) -> bool:
    if not model:
        return False
    return model.strip().lower().startswith("gpt-5")


def _error_is_empty(error: Any) -> bool:
    if error is None:
        return True
    if isinstance(error, dict):
        return len(error) == 0
    if isinstance(error, str):
        return error.strip() == ""
    return False


def _step_is_completed(step_entry: dict[str, Any]) -> bool:
    if not _error_is_empty(step_entry.get("error")):
        return False
    response_text = _ensure_response_text(step_entry) or ""
    return bool(response_text.strip())


def get_initial_context(sources_payload: dict[str, Any]) -> str:
    if not isinstance(sources_payload, dict):
        return ""
    sheets = sources_payload.get("sheets")
    if not isinstance(sheets, list):
        return ""
    for sheet_entry in sheets:
        if not isinstance(sheet_entry, dict):
            continue
        steps = sheet_entry.get("steps")
        if not isinstance(steps, list):
            continue
        for step_entry in steps:
            if not isinstance(step_entry, dict):
                continue
            step_number = step_entry.get("step_number")
            try:
                step_number = int(step_number)
            except (TypeError, ValueError):
                step_number = None
            if step_number != 1:
                continue
            text = _ensure_response_text(step_entry) or ""
            if not text:
                continue
            lines = text.splitlines()
            if lines and lines[0].lstrip().startswith("#"):
                lines = lines[1:]
            return "\n".join(lines).strip()
    return ""


def _find_step_anchor(ws, step_number: int, start_row: int = 1) -> Optional[Any]:
    pattern = re.compile(rf"^step\s*{step_number}(\D|$)", re.IGNORECASE)
    for cell in _iter_cells(ws):
        if cell.row < start_row:
            continue
        if isinstance(cell.value, str) and pattern.search(cell.value.strip()):
            return cell
    return None


def _parse_step_title(step_value: Any, step_number: int) -> tuple[str, str | None]:
    step_label = f"Step {step_number}"
    if not isinstance(step_value, str):
        return step_label, None
    raw = step_value.strip()
    if not raw:
        return step_label, None
    parts = re.split(r"\s[–\-:]\s", raw, maxsplit=1)
    if len(parts) > 1 and parts[1].strip():
        return step_label, parts[1].strip()
    return step_label, None


def _is_model_not_found_error(err: Exception) -> bool:
    msg = str(err).lower()
    if isinstance(err, NotFoundError):
        return ("model" in msg) or ("model_not_found" in msg)
    status_code = getattr(err, "status_code", None)
    if status_code == 404:
        return ("model" in msg) or ("model_not_found" in msg)
    return "model_not_found" in msg


def _format_error_reason(err: Exception) -> str:
    status_code = getattr(err, "status_code", None)
    if status_code is not None:
        return f"{status_code}: {err}"
    return str(err)


def _safe_write_text(path: Path, content: str) -> None:
    tmp_path = path.with_suffix(path.suffix + ".tmp")
    try:
        tmp_path.write_text(content, encoding="utf-8")
        tmp_path.replace(path)
    except OSError:
        return


def _safe_write_json(path: Path, payload: dict[str, Any]) -> None:
    try:
        content = json.dumps(payload, indent=2, ensure_ascii=False)
    except (TypeError, ValueError):
        return
    _safe_write_text(path, content)


def _safe_write_text_multi(paths: Iterable[Path], content: str) -> None:
    for path in paths:
        _safe_write_text(path, content)


def _safe_write_json_multi(paths: Iterable[Path], payload: dict[str, Any]) -> None:
    for path in paths:
        _safe_write_json(path, payload)


def _resolve_template_path() -> Path:
    template_path_value = os.getenv("PRIMER_WORD_TEMPLATE_PATH", "").strip()
    if not template_path_value:
        raise SystemExit(
            "ERROR: PRIMER_WORD_TEMPLATE_PATH is not set. DOCX output is required."
        )
    candidate = Path(template_path_value)
    if not candidate.is_absolute():
        repo_root = Path(__file__).resolve().parents[2]
        candidate = repo_root / candidate
    if not candidate.exists():
        raise SystemExit(f"ERROR: PRIMER_WORD_TEMPLATE_PATH not found: {candidate}")
    if candidate.is_dir():
        raise SystemExit(f"ERROR: PRIMER_WORD_TEMPLATE_PATH is a directory: {candidate}")
    try:
        with candidate.open("rb"):
            pass
    except OSError as err:
        raise SystemExit(
            f"ERROR: PRIMER_WORD_TEMPLATE_PATH is not readable: {candidate} ({err})"
        )
    return candidate


def _is_verbose() -> bool:
    for name in ("PRIMER_VERBOSE", "VERBOSE"):
        value = os.getenv(name, "").strip().lower()
        if value in ("1", "true", "yes", "y", "on"):
            return True
    return False


def _confirm_continue_after_timeout() -> bool:
    prompt = "Request timed out after 30 minutes. Continue and retry? [y/N]: "
    while True:
        try:
            answer = input(prompt).strip().lower()
        except EOFError:
            return False
        if answer in ("y", "yes"):
            return True
        if answer in ("n", "no", ""):
            return False
        print("Please answer y or n.")


def _call_openai_with_retries(
    client: OpenAI,
    request_kwargs: dict[str, Any],
    *,
    max_retries: int,
    base_sleep_seconds: float,
) -> Any:
    attempt = 0
    while True:
        try:
            return client.responses.create(**request_kwargs)
        except APITimeoutError:
            print(f"Request timed out after {int(_REQUEST_TIMEOUT_SECONDS / 60)} minutes.")
            if _confirm_continue_after_timeout():
                continue
            raise SystemExit("Aborted after timeout.")
        except RateLimitError as err:
            msg = str(err).lower()
            code = getattr(err, "code", None)
            should_retry = (code == "rate_limit_exceeded") or ("rate limit reached" in msg)
            if not should_retry:
                raise
            if attempt >= max_retries:
                raise
            match = re.search(r"try again in\s+(\d+)ms", msg)
            if match:
                sleep_seconds = (int(match.group(1)) + 50) / 1000.0
            else:
                sleep_seconds = min(base_sleep_seconds * (2**attempt), 10.0)
            print(
                f"Rate limited (attempt {attempt + 1}/{max_retries}). "
                f"Sleeping {sleep_seconds:.2f}s then retrying..."
            )
            time.sleep(sleep_seconds)
            attempt += 1


def generate_primer(
    output_dir: str | None = None,
    sheet: str | None = None,
    include: str | None = None,
    exclude: str | None = None,
    resume: bool = True,
    include_headings: bool | None = None,
    lead_input: str | None = None,
) -> None:
    env_path = find_dotenv(usecwd=True)
    load_dotenv(env_path, override=True)
    t0 = time.perf_counter()
    step = 0
    total_steps = 5

    def log_step(msg: str) -> None:
        nonlocal step
        step += 1
        print(f"[{step}/{total_steps}] {msg}", flush=True)

    include_headings_effective = (
        include_headings if include_headings is not None else get_include_headings(False)
    )

    lead_input_path = resolve_lead_input_path(lead_input)

    log_step("Loading lead_input.json")
    if not lead_input_path.exists():
        raise SystemExit(
            f"ERROR: lead_input.json not found at {lead_input_path}. "
            "Use --lead-input or set LEAD_INPUT_PATH."
        )
    lead = json.loads(lead_input_path.read_text(encoding="utf-8"))
    if not isinstance(lead, dict):
        raise SystemExit("ERROR: lead_input.json must contain a JSON object.")

    prompt_library_path = os.getenv("PROMPT_LIBRARY_PATH", "").strip()
    if not prompt_library_path:
        raise SystemExit("ERROR: PROMPT_LIBRARY_PATH is not set. Please set it in the .env file.")

    base_model = os.getenv("OPENAI_MODEL", "").strip() or "gpt-5.2"
    deep_model = os.getenv("OPENAI_DEEP_RESEARCH_MODEL", "").strip() or "o4-mini-deep-research"
    max_retries = int(os.getenv("OPENAI_MAX_RETRIES", "").strip() or 6)
    base_sleep_seconds = float(os.getenv("OPENAI_RETRY_BASE_SECONDS", "").strip() or 0.5)

    targets = resolve_output_targets(output_dir, lead)
    output_dir_path = targets["output_dir"]
    run_output_dir_path = targets["run_dir"]
    output_dirs: list[Path] = targets["output_dirs"]

    repo_root = targets["repo_root"]
    latest_dir = targets["latest_dir"]
    if repo_root is None:
        print(f"Resolved output dir: {output_dir_path}")
    else:
        print(f"Client repo: {repo_root}")
        print(f"Writing latest to: {latest_dir}")
        print(f"Writing run to: {run_output_dir_path}")

    company_name = _extract_company_name(lead)

    def write_output_text(filename: str, content: str) -> None:
        _safe_write_text_multi([path / filename for path in output_dirs], content)

    def write_output_json(filename: str, payload: dict[str, Any]) -> None:
        _safe_write_json_multi([path / filename for path in output_dirs], payload)
    lead_plus = dict(lead)
    lead_plus["client"] = company_name
    lead_plus["company"] = company_name
    lead_plus["company_name"] = company_name
    for key in list(lead_plus.keys()):
        lead_plus[str(key).upper()] = lead_plus[key]

    prompt_path = Path(prompt_library_path)
    if not prompt_path.is_absolute():
        repo_root = Path(__file__).resolve().parents[2]
        prompt_path = repo_root / prompt_path
    log_step("Loading prompt library (Excel)")
    workbook = load_workbook(prompt_path, data_only=True)

    def is_runnable_sheet(candidate_ws) -> bool:
        return (_find_anchor_exact(candidate_ws, "Instructions") is not None) and (
            _find_anchor_exact(candidate_ws, "Prompts") is not None
        )

    def parse_sheet_filter(filter_value: str, available: list[str]) -> set[str]:
        if "," in filter_value:
            requested = [item.strip() for item in filter_value.split(",") if item.strip()]
            available_by_lower = {name.lower(): name for name in available}
            matched = set()
            for name in requested:
                match = available_by_lower.get(name.lower())
                if match:
                    matched.add(match)
            return matched
        regex = re.compile(filter_value, re.IGNORECASE)
        return {name for name in available if regex.search(name)}

    runnable_sheets = [name for name in workbook.sheetnames if is_runnable_sheet(workbook[name])]

    if sheet:
        if sheet not in workbook.sheetnames:
            raise SystemExit(f"ERROR: sheet not found: {sheet}")
        if sheet not in runnable_sheets:
            raise SystemExit(
                f"ERROR: sheet is not runnable (missing Instructions/Prompts): {sheet}"
            )
        selected_sheets = [sheet]
    else:
        selected_sheets = list(runnable_sheets)

    if sheet is None:
        if include:
            selected_sheets = [
                name for name in selected_sheets if name in parse_sheet_filter(include, selected_sheets)
            ]
        if exclude:
            excluded = parse_sheet_filter(exclude, selected_sheets)
            selected_sheets = [name for name in selected_sheets if name not in excluded]

    if not selected_sheets:
        raise SystemExit("ERROR: no runnable sheets selected.")

    print(f"Sheets to run: {', '.join(selected_sheets)}")

    primer_sections: list[str] = ["# Commercial Primer"] if include_headings_effective else []
    sources_path = output_dir_path / "sources.json"
    sources_payload: dict[str, Any] = {
        "prompt_library_path": str(prompt_path),
        "sheets": [],
    }
    if resume and sources_path.exists():
        try:
            loaded_payload = json.loads(sources_path.read_text(encoding="utf-8"))
        except (OSError, ValueError):
            loaded_payload = None
        if isinstance(loaded_payload, dict):
            sources_payload = _sanitize_sources_payload(loaded_payload)
            sources_payload["prompt_library_path"] = str(prompt_path)

    if isinstance(sources_payload.get("sheets"), list):
        existing_sheets = set(workbook.sheetnames)
        original_count = len(sources_payload["sheets"])
        sources_payload["sheets"] = [
            entry
            for entry in sources_payload["sheets"]
            if isinstance(entry, dict)
            and isinstance(entry.get("name"), str)
            and entry["name"] in existing_sheets
        ]
        removed_count = original_count - len(sources_payload["sheets"])
        if removed_count:
            print(
                f"Removed {removed_count} sheet(s) from sources.json not present in prompt library."
            )

    prev_sheet_output_text = get_initial_context(sources_payload)
    prev_sheet_name: str | None = "seed:intro" if prev_sheet_output_text.strip() else None
    client = OpenAI(timeout=_REQUEST_TIMEOUT_SECONDS)

    sheets_by_name: dict[str, dict[str, Any]] = {}
    for sheet_entry in sources_payload.get("sheets", []) or []:
        if isinstance(sheet_entry, dict) and isinstance(sheet_entry.get("name"), str):
            sheets_by_name[sheet_entry["name"]] = sheet_entry

    for sheet_index, sheet_name in enumerate(selected_sheets, start=1):
        ws = workbook[sheet_name]
        log_step(f"Parsing sheet anchors and building prompt ({sheet_name})")
        instructions_cell = _require_anchor(_find_anchor_exact(ws, "Instructions"), "Instructions")
        web_search_cell = _require_anchor(
            _find_anchor_exact(ws, "Web Search", start_row=instructions_cell.row),
            "Web Search",
        )
        web_search_value = _first_right_value(ws, web_search_cell.row, web_search_cell.column)
        if web_search_value is None:
            raise SystemExit("ERROR: missing anchor: Web Search")
        deep_research_cell = _require_anchor(
            _find_anchor_exact(ws, "Deep Research", start_row=instructions_cell.row),
            "Deep Research",
        )
        deep_research_value = _first_right_value(
            ws, deep_research_cell.row, deep_research_cell.column, limit=20
        )
        if deep_research_value is None:
            raise SystemExit("ERROR: missing anchor: Deep Research")

        reasoning_effort: str | None = None
        gpt_model_cell = _find_anchor_exact(ws, "GPT Model", start_row=instructions_cell.row)
        if gpt_model_cell is not None:
            gpt_model_value = _first_right_value(ws, gpt_model_cell.row, gpt_model_cell.column)
            if isinstance(gpt_model_value, str):
                normalized = gpt_model_value.strip()
                if normalized == "Thinking - Reasoning: Extended":
                    reasoning_effort = "xhigh"
                elif normalized == "Thinking":
                    reasoning_effort = "high"
                elif normalized == "Auto":
                    reasoning_effort = None

        web_search_enabled = _normalize(str(web_search_value)).startswith("enable")
        deep_research_requested = _normalize(str(deep_research_value)).startswith("enable")
        if deep_research_requested and not web_search_enabled:
            raise SystemExit(
                "ERROR: Deep Research is Enable but Web Search is not Enable. Deep Research needs a data source."
            )

        sheet_entry = sheets_by_name.get(sheet_name)
        if sheet_entry is None:
            sheet_entry = {
                "name": sheet_name,
                "web_search": web_search_enabled,
                "deep_research_requested": deep_research_requested,
                "deep_research_effective": False,
                "deep_research_error_reason": None,
                "steps": [],
            }
            sources_payload.setdefault("sheets", [])
            sources_payload["sheets"].append(sheet_entry)
            sheets_by_name[sheet_name] = sheet_entry
        else:
            sheet_entry["web_search"] = web_search_enabled
            sheet_entry["deep_research_requested"] = deep_research_requested
            sheet_entry.setdefault("steps", [])

        if include_headings_effective:
            primer_sections.append(f"## {sheet_name}")

        prompts_cell = _require_anchor(_find_anchor_exact(ws, "Prompts"), "Prompts")
        current_sheet_output_sections: list[str] = []
        steps_by_number: dict[int, dict[str, Any]] = {}
        for existing_step in sheet_entry.get("steps", []) or []:
            if not isinstance(existing_step, dict):
                continue
            raw_number = existing_step.get("step_number")
            if isinstance(raw_number, int):
                steps_by_number[raw_number] = existing_step
            else:
                try:
                    step_num = int(raw_number)
                except (TypeError, ValueError):
                    continue
                steps_by_number[step_num] = existing_step
        step_number = 1
        while True:
            step_cell = _find_step_anchor(ws, step_number, start_row=prompts_cell.row)
            if step_cell is None:
                break
            step_label, step_title_clean = _parse_step_title(step_cell.value, step_number)
            if step_title_clean:
                heading = f"### {step_label} — {step_title_clean}"
            else:
                heading = f"### {step_label}"

            existing_step_entry = steps_by_number.get(step_number)
            if existing_step_entry is None:
                step_entry = {
                    "step_number": step_number,
                    "title": step_title_clean or step_label,
                    "prompt": None,
                    "response_text": "",
                    "model": None,
                    "reasoning_effort_requested": None,
                    "reasoning_effort_effective": None,
                    "web_search": web_search_enabled,
                    "deep_research_requested": deep_research_requested,
                    "web_tool_type": None,
                    "deep_research_effective": False,
                    "deep_research_error_reason": None,
                    "error": None,
                }
                sheet_entry["steps"].append(step_entry)
                steps_by_number[step_number] = step_entry
            else:
                step_entry = existing_step_entry
                step_entry.setdefault("title", step_title_clean or step_label)
                step_entry.setdefault("prompt", None)
                step_entry.setdefault("response_text", "")
                step_entry.setdefault("model", None)
                step_entry.setdefault("reasoning_effort_requested", None)
                step_entry.setdefault("reasoning_effort_effective", None)
                step_entry.setdefault("web_search", web_search_enabled)
                step_entry.setdefault("deep_research_requested", deep_research_requested)
                step_entry.setdefault("web_tool_type", None)
                step_entry.setdefault("deep_research_effective", False)
                step_entry.setdefault("deep_research_error_reason", None)
                step_entry.setdefault("error", None)
                if "citations" in step_entry and not isinstance(step_entry["citations"], list):
                    step_entry["citations"] = []
            try:
                next_step_cell = _find_step_anchor(ws, step_number + 1, start_row=step_cell.row + 1)
                end_row = next_step_cell.row if next_step_cell is not None else ws.max_row + 1
                suggested_prompt_cell = _find_anchor_exact_in_window(
                    ws, "Suggested Prompt", start_row=step_cell.row, end_row=end_row
                )
                if suggested_prompt_cell is None:
                    raise ValueError(f"ERROR: missing anchor: Suggested Prompt (step {step_number})")
                suggested_prompt_value = _first_right_value(
                    ws, suggested_prompt_cell.row, suggested_prompt_cell.column, limit=20
                )
                if not isinstance(suggested_prompt_value, str) or not suggested_prompt_value.strip():
                    raise ValueError("ERROR: missing anchor: Suggested Prompt")

                prompt_original = _replace_placeholders(suggested_prompt_value, lead_plus)
                context_text = prev_sheet_output_text
                if current_sheet_output_sections:
                    sheet_context = "\n\n".join(current_sheet_output_sections).strip()
                    if sheet_context:
                        if context_text:
                            context_text = f"{context_text}\n\n{sheet_context}"
                        else:
                            context_text = sheet_context
                (
                    prompt_final,
                    injected_prev_sheet_context,
                    prev_sheet_name_used,
                    prev_sheet_output_chars,
                ) = _post_process_prompt(prompt_original, prev_sheet_name, context_text)
                step_entry["title"] = step_title_clean or step_label
                step_entry["prompt"] = prompt_final
                step_entry["web_search"] = web_search_enabled
                step_entry["deep_research_requested"] = deep_research_requested

                output_text = ""
                error_message: str | None = None
                call_label = f"[sheet {sheet_index}/{len(selected_sheets)}][step {step_number}]"

                if resume and existing_step_entry is not None and _step_is_completed(existing_step_entry):
                    output_text = _ensure_response_text(step_entry) or ""
                    if output_text:
                        step_entry["response_text"] = output_text.strip()
                    elif step_entry.get("response_text") is None:
                        step_entry["response_text"] = ""
                    print(f"{call_label} SKIP completed step")
                else:
                    model = base_model
                    web_tool_type: str | None = "web_search" if web_search_enabled else None
                    deep_research_effective = False
                    deep_research_error_reason: str | None = None
                    response: Any | None = None

                    call_start = time.perf_counter()
                    print(f"{call_label} Calling OpenAI...")
                    if deep_research_requested:
                        request_kwargs = {"model": deep_model, "input": prompt_final}
                        if web_search_enabled:
                            request_kwargs["tools"] = [{"type": "web_search_preview"}]
                        try:
                            response = _call_openai_with_retries(
                                client,
                                request_kwargs,
                                max_retries=max_retries,
                                base_sleep_seconds=base_sleep_seconds,
                            )
                            model = deep_model
                            web_tool_type = "web_search_preview" if web_search_enabled else None
                            deep_research_effective = True
                        except (RateLimitError, NotFoundError) as err:
                            if isinstance(err, RateLimitError):
                                deep_research_error_reason = _format_error_reason(err)
                            elif not _is_model_not_found_error(err):
                                raise
                            else:
                                deep_research_error_reason = _format_error_reason(err)
                            model = base_model
                            web_tool_type = "web_search" if web_search_enabled else None
                            request_kwargs = {"model": model, "input": prompt_final}
                            if web_search_enabled:
                                request_kwargs["tools"] = [{"type": "web_search"}]
                            if reasoning_effort is not None and _model_supports_reasoning_effort(model):
                                request_kwargs["reasoning"] = {"effort": reasoning_effort}
                            try:
                                response = _call_openai_with_retries(
                                    client,
                                    request_kwargs,
                                    max_retries=max_retries,
                                    base_sleep_seconds=base_sleep_seconds,
                                )
                            except RateLimitError as err_fallback:
                                error_message = f"{type(err_fallback).__name__}: {err_fallback}"
                    else:
                        request_kwargs = {"model": model, "input": prompt_final}
                        if web_search_enabled:
                            request_kwargs["tools"] = [{"type": "web_search"}]
                        if reasoning_effort is not None and _model_supports_reasoning_effort(model):
                            request_kwargs["reasoning"] = {"effort": reasoning_effort}
                        try:
                            response = _call_openai_with_retries(
                                client,
                                request_kwargs,
                                max_retries=max_retries,
                                base_sleep_seconds=base_sleep_seconds,
                            )
                        except RateLimitError as err:
                            error_message = f"{type(err).__name__}: {err}"

                    effort_effective = None if deep_research_effective else reasoning_effort
                    call_elapsed = time.perf_counter() - call_start
                    print(f"{call_label} Done in {format_seconds(call_elapsed)}")

                    print(
                        " ".join(
                            [
                                f"model={model}",
                                f"effort={effort_effective}",
                                f"web_search={web_search_enabled}",
                                f"deep_research_requested={deep_research_requested}",
                                f"deep_research_effective={deep_research_effective}",
                                f"web_tool_type={web_tool_type}",
                            ]
                        )
                    )

                    if response is not None:
                        output_text = _extract_output_text_from_response(response) or ""
                    citations: list[str] = []
                    if response is not None:
                        citations = _extract_citations_from_response(response, output_text)
                    if not output_text and error_message is None:
                        error_message = "No output returned."

                    step_entry["model"] = model
                    step_entry["reasoning_effort_requested"] = reasoning_effort
                    step_entry["reasoning_effort_effective"] = effort_effective
                    step_entry["web_search"] = web_search_enabled
                    step_entry["deep_research_requested"] = deep_research_requested
                    step_entry["web_tool_type"] = web_tool_type
                    step_entry["deep_research_effective"] = deep_research_effective
                    step_entry["deep_research_error_reason"] = deep_research_error_reason
                    step_entry["error"] = error_message
                    step_entry["response_text"] = output_text.strip() if output_text else ""
                    if citations:
                        step_entry["citations"] = citations
                    elif "citations" in step_entry:
                        step_entry["citations"] = []

                    if deep_research_effective:
                        sheet_entry["deep_research_effective"] = True
                    if deep_research_error_reason and sheet_entry["deep_research_error_reason"] is None:
                        sheet_entry["deep_research_error_reason"] = deep_research_error_reason


                if include_headings_effective:
                    primer_sections.append(heading)
                if output_text:
                    trimmed_output = output_text.strip()
                    primer_sections.append(trimmed_output)
                    current_sheet_output_sections.append(trimmed_output)
                elif include_headings_effective:
                    if error_message:
                        primer_sections.append(f"Error: {error_message}")
                    else:
                        primer_sections.append("Error: No output returned.")
            except Exception as err:
                step_entry["error"] = f"{type(err).__name__}: {err}"
                step_entry["response_text"] = ""
                if include_headings_effective:
                    primer_sections.append(heading)
                    primer_sections.append(f"Error: {err}")

            primer_content = "\n\n".join(primer_sections).strip() + "\n"
            write_output_text("primer.md", primer_content)
            write_output_json("sources.json", sources_payload)
            step_number += 1

        prev_sheet_output_text = "\n\n".join(current_sheet_output_sections).strip()
        prev_sheet_name = sheet_name

    if not sources_payload["sheets"]:
        raise SystemExit("ERROR: no runnable sheets executed.")

    log_step("Saving outputs")
    primer_content = "\n\n".join(primer_sections).strip() + "\n"
    write_output_text("primer.md", primer_content)
    write_output_json("sources.json", sources_payload)

    print(f"Saved: {output_dir_path / 'primer.md'}")
    print(f"Saved: {output_dir_path / 'sources.json'}")
    if run_output_dir_path is not None:
        print(f"Saved: {run_output_dir_path / 'primer.md'}")
        print(f"Saved: {run_output_dir_path / 'sources.json'}")

    template_path = _resolve_template_path()
    try:
        from primer_ops.render_docx import render_primer_docx
    except Exception as err:
        print(f"ERROR: DOCX renderer import failed: {err}")
        if _is_verbose():
            traceback.print_exc()
        raise SystemExit(1)

    docx_errors: list[str] = []
    for out_dir in output_dirs:
        md_path = out_dir / "primer.md"
        if not md_path.exists():
            msg = f"{md_path} not found"
            print(f"ERROR: DOCX render failed: {msg}")
            docx_errors.append(msg)
            continue
        docx_path = md_path.with_suffix(".docx")
        try:
            render_primer_docx(
                str(md_path),
                str(docx_path),
                str(template_path),
            )
            print(f"Saved: {docx_path}")
        except Exception as err:
            msg = f"{docx_path}: {err}"
            print(f"ERROR: DOCX render failed: {err}")
            if _is_verbose():
                traceback.print_exc()
            docx_errors.append(msg)

    if docx_errors:
        raise SystemExit("ERROR: DOCX render failed (see above).")
    elapsed = time.perf_counter() - t0
    print(f"Done in {format_seconds(elapsed)}", flush=True)
